# -*- coding: utf-8 -*-
"""
Created on Wed Nov  3 12:33:37 2021

@author: Sriram Prasad S
"""


import sys
import calend
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from PyQt5.QtWidgets import QLabel, QLineEdit, QComboBox
from PyQt5.QtWidgets import QWidget, QDialog
from PyQt5.QtWidgets import QFormLayout, QVBoxLayout, QHBoxLayout
from PyQt5.QtWidgets import QPushButton, QSpinBox
from PyQt5.QtGui import QFont
from pathlib import Path



class DataEntry(QWidget):
    def __init__(self):
        super().__init__()
        
        
        self.setGeometry(900, 100, 500, 500)
        self.setWindowTitle('Data Entry Settings')
        layout2 = QFormLayout()
        
        self.dset = QLineEdit()
        self.dset.setFont(QFont('Times',10))
        laylb1 = QLabel('Date: ')
        laylb1.setFont(QFont('Times',10))
        button1 = QPushButton('Set date')
        button1.setFont(QFont('Times',10))
        
        self.shif = QComboBox()
        self.shif.addItems(['DAY','NIGHT'])
        self.shif.setFont(QFont('Times',10))
        laylb2 = QLabel('Shift: ')
        laylb2.setFont(QFont('Times',10))
        
        
        self.mcno = QComboBox()
        self.mcno.addItems(['1','2','3'])
        self.mcno.setFont(QFont('Times',10))
        laylb3 = QLabel('Machine No.: ')
        laylb3.setFont(QFont('Times',10))
        
        laylb4 = QLabel('Hours run: ')
        laylb4.setFont(QFont('Times',10))
        laylb5 = QLabel('minutes run: ')
        laylb5.setFont(QFont('Times',10))
        
        
        
        
        
        """ box layout for hrs and minutes (hrs run)"""
        
        laylb4 = QLabel('Hours run: ')
        laylb4.setFont(QFont('Times',10))
        
        
        hbox = QHBoxLayout()
        self.hrrun = QSpinBox()
        self.hrrun.setRange(0,12)
        self.mnrun = QSpinBox()
        self.mnrun.setRange(0,60)
        self.hrrun.setFont(QFont('Times',10))
        self.mnrun.setFont(QFont('Times',10))
        hbox.addWidget(self.hrrun)
        
        hbox.addWidget(self.mnrun)
        
        cont1 = QWidget()
        cont1.setLayout(hbox)
        """ ending box layout for hrs and min (hrs run) """
        
        
        
        
        
        
        
        
        """ box layout for hrs and minutes (breakdown)"""
        
        laylb5 = QLabel('Breakdown : ')
        laylb5.setFont(QFont('Times',10))
        
        
        hbox2 = QHBoxLayout()
        self.hrbre = QSpinBox()
        self.hrbre.setRange(0,12)
        self.mnbre = QSpinBox()
        self.mnbre.setRange(0,60)
        self.hrbre.setFont(QFont('Times',10))
        self.mnbre.setFont(QFont('Times',10))
        hbox2.addWidget(self.hrbre)
        
        hbox2.addWidget(self.mnbre)
        
        cont2 = QWidget()
        cont2.setLayout(hbox2)
        
        """ ending box layout for hrs and min (breakdown) """
        
        
        
        
        
        
        
        
        """ box layout for hrs and minutes (power failure)"""
        
        laylb6 = QLabel('Power Failure : ')
        laylb6.setFont(QFont('Times',10))
        
        
        hbox3 = QHBoxLayout()
        self.hrpow = QSpinBox()
        self.hrpow.setRange(0,12)
        self.mnpow = QSpinBox()
        self.mnpow.setRange(0,60)
        self.hrpow.setFont(QFont('Times',10))
        self.mnpow.setFont(QFont('Times',10))
        hbox3.addWidget(self.hrpow)
        
        hbox3.addWidget(self.mnpow)
        
        cont3 = QWidget()
        cont3.setLayout(hbox3)
        
        """ ending box layout for hrs and min (power failure) """
        
        
        
        
        
        
        
        """ box layout for hrs and minutes (maintenance)"""
        
        laylb7 = QLabel('Maintenance : ')
        laylb7.setFont(QFont('Times',10))
        
        
        hbox4 = QHBoxLayout()
        self.hrmai = QSpinBox()
        self.hrmai.setRange(0,12)
        self.mnmai = QSpinBox()
        self.mnmai.setRange(0,60)
        self.hrmai.setFont(QFont('Times',10))
        self.mnmai.setFont(QFont('Times',10))
        
        hbox4.addWidget(self.hrmai)
        hbox4.addWidget(self.mnmai)
        
        cont4 = QWidget()
        cont4.setLayout(hbox4)
        
        """ ending box layout for hrs and min (maintenance) """
        
        """
        idle spindles 
        
        """
        
        
        self.idSpin = QSpinBox()
        self.idSpin.setRange(0,432)
        laylb8 = QLabel('No. of idle Spindles: ')
        laylb8.setFont(QFont('Times',10))
        self.idSpin.setFont(QFont('Times',10))
        
        
        
        
        
        
        
        button2 = QPushButton('Save to file')
        button2.setFont(QFont('Times',10))
        
        
        
        layout2.addRow(laylb1, self.dset)
        layout2.addWidget(button1)
        layout2.addRow(laylb2, self.shif)
        layout2.addRow(laylb3,self.mcno)
        layout2.addRow(laylb8,self.idSpin)
        layout2.addRow(laylb4,cont1)
        layout2.addRow(laylb5,cont2)
        layout2.addRow(laylb6,cont3)
        layout2.addRow(laylb7,cont4)
        layout2.addWidget(button2)
        
        button1.clicked.connect(self.clickCal)
        button2.clicked.connect(self.clickSaveFile)
        
        self.setLayout(layout2)
        
    def clickCal(self):
        self.cal1 = calend.Window()
        self.cal1.show()
        self.cal1.calendar.clicked.connect(self.calChange)
        
    
    def calChange(self):
        dateselected = self.cal1.calendar.selectedDate()
        date_in_string = dateselected.toPyDate()
        date_in_string = str(date_in_string.strftime('%d/%m/%Y'))
        self.dset.setText(date_in_string)
        self.cal1.close()
        
    
        
    def clickSaveFile(self):
        
        f1 = open("mset.txt","r")
        lines = f1.readlines()
        
        count = float(lines[0]) 
        mpm = float(lines[3])
        f1.close()
        
        prod = (mpm * 0.0354 * (432-float(self.idSpin.value())) * (float(self.hrrun.value()) + (float(self.mnrun.value())/60)) / count)
        prod = float("{:.2f}".format(prod))
        
        date = self.dset.text()
        shift = self.shif.currentText()
        machno = self.mcno.currentText()
        idlespin = self.idSpin.text()
        idlespin = int(idlespin)
        
        machno = int(machno)
        
        hrsrun = float(self.hrrun.value()) + (float(self.mnrun.value())/60)
        hrsrun = float("{:.2f}".format(hrsrun))
        
        
        brkdown = float(self.hrbre.value()) + (float(self.mnbre.value())/60)
        brkdown = float("{:.2f}".format(brkdown))
        
        
        
        powerfail = float(self.hrpow.value()) + (float(self.mnpow.value())/60)
        powerfail = float("{:.2f}".format(powerfail))
        
    
        maintenance = float(self.hrmai.value()) + (float(self.mnmai.value())/60)
        maintenance = float("{:.2f}".format(maintenance))
        
        
        efficiency = ((432-idlespin) * hrsrun )/ (432 * 12)
        efficiency = float("{:.2f}".format(efficiency))
        self.couxy = count
        
        path_to_file = "report.xlsx"
        path = Path(path_to_file)

        if path.is_file():
            wb = openpyxl.load_workbook(path_to_file)
            sheet = wb.active
            sheet.append([date,shift,count,machno,idlespin,hrsrun,prod,brkdown,powerfail,maintenance,efficiency])
            wb.save(path_to_file)
            wb.close()
            
        
        else:
            self.initialSave()
            wb = openpyxl.load_workbook(path_to_file)
            sheet = wb.active
            sheet.append([date,shift,count,machno,idlespin,hrsrun,prod,brkdown,powerfail,maintenance,efficiency])
            wb.save(path_to_file)
            wb.close()
            
        
        self.sizeIncrease()
        self.widthIncrease()
        self.dialogSave()
        
    
    
    def initialSave(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
         
        headers = ['DATE','SHIFT','COUNT','MACHINE NO','NO. OF IDLE SPINDLES','HRS RUN','TOTAL PROD.', 'BREAKDOWN','POWER FAILURE','MAINTENANCE','EFFICIENCY']
        sheet.append(headers)
        
        wb.save("report.xlsx")
        wb.close()
        
           
    def dialogSave(self):
        mydialog = QDialog(self)
        mydialog.setGeometry(700, 100, 400, 150)
        mydialog.setWindowTitle('Updated')
        
        layout = QVBoxLayout()
        label1 = QLabel("Shift :"+ self.shif.currentText())
        label2 = QLabel("Machine no: " + self.mcno.currentText())
        label3 = QLabel("Count :" + str(self.couxy) )
        
        label1.setFont(QFont('Times',10))
        label2.setFont(QFont('Times',10))
        label3.setFont(QFont('Times',10))
        
        
        layout.addWidget(label1)
        layout.addWidget(label2)
        layout.addWidget(label3)
        mydialog.setLayout(layout)
        
        mydialog.show()
        mydialog.exec_()
            
    
    def sizeIncrease(self):
        wb = openpyxl.load_workbook("report.xlsx")
        sheet = wb.active
        
       
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    cell.font = Font(size = 12)
        wb.save("report.xlsx")
        
        wb.close()
        
    def widthIncrease(self):
        wb = openpyxl.load_workbook("report.xlsx")
        sheet = wb.active
        
        varalph = ['A','B','C','D','E','F','G','H','I','J','K']
        
        for x in varalph:
            sheet.column_dimensions[x].width = 15
        wb.save("report.xlsx")
        
        wb.close()
    
        
 